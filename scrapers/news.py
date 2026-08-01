#!/usr/bin/env python3
"""
FSIS News Feed scraper.

Runs hourly from .github/workflows/news-feed.yml:
    python scrapers/news.py docs/data/recalls.xlsx

Pulls recent food-pathogen news from a curated set of RSS feeds, filters by
the VisiPilot pathogen whitelist, dedupes against existing NEWS rows by Link, and
enforces a 7-day rolling retention (rows with Published > 7 days ago are
dropped on every run — per the VisiPilot FSIS spec).

Writes back only the NEWS sheet. Recalls and Pending sheets are untouched.
"""
from __future__ import annotations

import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import feedparser
import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

NEWS_HEADERS = ["Published (UTC)", "Pathogen", "Event", "Source",
                "Title", "Link", "Retrieved (UTC)", "Category"]

RETENTION_DAYS = 7
HTTP_TIMEOUT = 20
USER_AGENT = "VisiPilot-FSIS-NewsFeed/1.0 (+https://visipilot.com)"

# Curated RSS/Atom feeds — food-safety focused outlets. Tagged Category=
# "Outbreak" and REQUIRE a pathogen-whitelist match to be kept (see
# PATHOGEN_PATTERNS below) — this group is specifically outbreak/illness
# news, not general food-industry coverage.
FEEDS = [
    ("Food Safety News",        "https://www.foodsafetynews.com/feed/"),
    ("Food Poisoning Bulletin", "https://foodpoisoningbulletin.com/feed/"),
    ("Outbreak News Today",     "https://outbreaknewstoday.com/feed/"),
    ("CIDRAP",                  "https://www.cidrap.umn.edu/news-perspective/topic/foodborne-disease/feed"),
    ("Food Safety Magazine",    "https://www.food-safety.com/rss/topic/296-food-safety"),
    ("Barfblog",                "https://barfblog.com/feed/"),
]

# Official regulator/industry announcement feeds — Category="Regulatory".
# These cover a much broader spectrum than outbreak news (labeling,
# additives, risk assessments, policy) so they do NOT require a pathogen
# match to be kept — only the same noise/exclusion guard as FEEDS.
OFFICIAL_FEEDS = [
    ("EFSA",               "https://www.efsa.europa.eu/en/all/rss"),
    ("FDA Press Releases", "https://www.fda.gov/about-fda/contact-fda/stay-informed/rss-feeds/press-releases/rss.xml"),
    ("Food Navigator",     "https://www.foodnavigator.com/arc/outboundfeeds/rss/"),
    ("FSA UK News",        "https://www.food.gov.uk/rss-feed/news"),
    ("ANSES",              "https://www.anses.fr/fr/rss.xml"),
    ("DGCCRF",              "https://www.economie.gouv.fr/dgccrf/rss"),
]

# Pathogen whitelist — first regex to match wins. Order matters: put more
# specific patterns before generic ones. Matches are run against title+summary.
# Keep in sync with the FSIS pathogen whitelist in the Apps Script.
PATHOGEN_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("E. coli / STEC",        re.compile(r"\b(e\.?\s*coli|stec|o157|shiga[- ]?toxin)\b", re.I)),
    ("Listeria monocytogenes",re.compile(r"\blisteria\b",                                re.I)),
    ("Salmonella spp.",       re.compile(r"\bsalmonella\b",                              re.I)),
    ("Clostridium botulinum", re.compile(r"\b(botulinum|botulism)\b",                    re.I)),
    ("Norovirus",             re.compile(r"\bnorovirus\b",                               re.I)),
    ("Hepatitis A",           re.compile(r"\bhepatitis\s*a\b",                           re.I)),
    ("Campylobacter",         re.compile(r"\bcampylobacter\b",                           re.I)),
    ("Vibrio",                re.compile(r"\bvibrio\b",                                  re.I)),
    ("Cyclospora",            re.compile(r"\bcyclospora(?:iasis)?\b",                    re.I)),
    ("Yersinia",              re.compile(r"\byersinia\b",                                re.I)),
    ("Bacillus cereus",       re.compile(r"\bbacillus\s*cereus\b",                       re.I)),
    ("Brucella",              re.compile(r"\bbrucell(a|osis)\b",                         re.I)),
    ("Aflatoxin",             re.compile(r"\baflatoxin\b",                               re.I)),
    ("Mycotoxin",             re.compile(r"\bmycotoxin|ochratoxin|fumonisin\b",          re.I)),
]

# Event classification — order matters (outbreak beats recall beats illness).
EVENT_RULES: list[tuple[str, re.Pattern]] = [
    ("Outbreak", re.compile(r"\b(outbreak|cluster|linked\s+to|sickened)\b", re.I)),
    ("Recall",   re.compile(r"\brecall(ed|ing|s)?\b",                      re.I)),
    ("Illness",  re.compile(r"\b(illness(es)?|hospitali[sz]ed|death|died|fatal)\b", re.I)),
]

# Noise / exclusions — titles containing these are skipped even if a pathogen
# matches (label errors, allergen-only recalls, foreign material — per FSIS
# whitelist rules).
EXCLUDE_PATTERNS = [
    re.compile(r"\ballerg(en|ic|y)\b",           re.I),
    re.compile(r"\bundeclared\s+(milk|egg|soy|wheat|nut|peanut|sesame|sulfite)\b", re.I),
    re.compile(r"\bforeign\s+(material|object|body)\b", re.I),
    re.compile(r"\bmetal\s+fragment",            re.I),
    re.compile(r"\bplastic\s+fragment",          re.I),
    re.compile(r"\bmislabel",                    re.I),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now_utc() -> datetime:
    return datetime.now(timezone.utc)

def _fmt_utc(dt: datetime) -> str:
    """Format as '2026-04-17 05:37 UTC' — matches existing NEWS sheet rows."""
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

def _parse_utc(s: str) -> datetime | None:
    """Parse the 'YYYY-MM-DD HH:MM UTC' format back to a datetime."""
    if not s:
        return None
    s = str(s).strip().replace(" UTC", "")
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None

def _entry_published(entry) -> datetime | None:
    """Extract UTC-aware datetime from a feedparser entry."""
    for key in ("published_parsed", "updated_parsed", "created_parsed"):
        t = getattr(entry, key, None) or entry.get(key) if hasattr(entry, "get") else None
        if t:
            try:
                return datetime(*t[:6], tzinfo=timezone.utc)
            except (TypeError, ValueError):
                pass
    return None

def _classify_pathogen(text: str) -> str | None:
    for name, pat in PATHOGEN_PATTERNS:
        if pat.search(text):
            return name
    return None

def _classify_event(text: str) -> str:
    for name, pat in EVENT_RULES:
        if pat.search(text):
            return name
    return "News"

def _is_excluded(text: str) -> bool:
    return any(p.search(text) for p in EXCLUDE_PATTERNS)

def _normalize_link(url: str) -> str:
    """Strip tracking params and trailing slashes for stable dedup."""
    if not url:
        return ""
    url = url.strip()
    # Drop common trackers
    url = re.sub(r"[?&](utm_[^=]+|fbclid|gclid|mc_cid|mc_eid|ref)=[^&]*", "", url)
    url = re.sub(r"[?&]+$", "", url).rstrip("/")
    return url

def _fetch_feed(url: str) -> list:
    """Fetch a feed with a real User-Agent (some hosts 403 the default UA)."""
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        parsed = feedparser.parse(r.content)
        return list(parsed.entries or [])
    except Exception as e:
        print(f"  ! feed failed: {url}  ({e})", file=sys.stderr)
        return []

def _clean_summary(raw: str) -> str:
    """Strip script/style/comment content and HTML tags, decode common
    entities, and belt-and-braces-guard against JS leaking into plain-text
    feeds (observed in Food Safety News RSS, audit 2026-04-28)."""
    summary = raw or ""
    summary = re.sub(r"<script\b[^>]*>.*?</script\s*>", " ", summary,
                     flags=re.DOTALL | re.IGNORECASE)
    summary = re.sub(r"<style\b[^>]*>.*?</style\s*>", " ", summary,
                     flags=re.DOTALL | re.IGNORECASE)
    summary = re.sub(r"<!--.*?-->", " ", summary, flags=re.DOTALL)
    summary = re.sub(r"<[^>]+>", " ", summary)
    summary = (summary
               .replace("&nbsp;", " ")
               .replace("&amp;", "&")
               .replace("&lt;", "<")
               .replace("&gt;", ">")
               .replace("&quot;", '"')
               .replace("&#39;", "'")
               .replace("&apos;", "'"))
    for marker in ("{socials", "window.socials", "window.adsbygoogle",
                   "document.cookie", "addEventListener("):
        idx = summary.find(marker)
        if idx >= 0:
            summary = summary[:idx].rstrip()
            break
    return re.sub(r"\s+", " ", summary).strip()

def _process_feed_group(feeds, cutoff, seen_links, retrieved, require_pathogen, category):
    """Fetch + filter one group of feeds (FEEDS or OFFICIAL_FEEDS), returning
    (new_rows, per_feed_added). Mutates seen_links as it goes so the two
    groups never produce duplicate links against each other."""
    new_rows = []
    per_feed_added: dict[str, int] = {}
    for source, url in feeds:
        entries = _fetch_feed(url)
        print(f"[news] {source}: {len(entries)} entries")
        for e in entries:
            title = (e.get("title") or "").strip()
            link  = _normalize_link(e.get("link") or "")
            if not title or not link:
                continue
            if link in seen_links:
                continue
            pub_dt = _entry_published(e)
            if not pub_dt or pub_dt < cutoff:
                continue
            summary = _clean_summary(e.get("summary") or e.get("description") or "")
            for suffix in (f"| {source}", "| Food Safety News", "| CIDRAP",
                           "| Outbreak News Today", "| Reuters", "| BBC News",
                           "| The Guardian"):
                if title.endswith(suffix):
                    title = title[:-len(suffix)].rstrip(" -|")
                    break
            text = f"{title} {summary}"
            if _is_excluded(text):
                continue
            pathogen = _classify_pathogen(text)
            if require_pathogen and not pathogen:
                continue
            event = _classify_event(text)
            new_rows.append([
                _fmt_utc(pub_dt), pathogen or "", event, source, title, link,
                retrieved, category,
            ])
            seen_links.add(link)
            per_feed_added[source] = per_feed_added.get(source, 0) + 1
        time.sleep(0.5)
    return new_rows, per_feed_added

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if len(sys.argv) < 2:
        print("usage: news.py <path/to/recalls.xlsx>", file=sys.stderr)
        return 2
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    now = _now_utc()
    cutoff = now - timedelta(days=RETENTION_DAYS)
    retrieved = _fmt_utc(now)

    print(f"[news] loading {xlsx_path}")
    wb = load_workbook(xlsx_path)
    if "NEWS" in wb.sheetnames:
        ws = wb["NEWS"]
    else:
        ws = wb.create_sheet("NEWS")
        ws.append(NEWS_HEADERS)

    # Read existing rows, apply 7-day retention, build dedup set.
    existing: list[list] = []
    seen_links: set[str] = set()
    dropped_old = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not any(row):
            continue
        row = list(row) + [""] * (len(NEWS_HEADERS) - len(row))
        row = row[:len(NEWS_HEADERS)]
        pub_dt = _parse_utc(row[0])
        if pub_dt and pub_dt < cutoff:
            dropped_old += 1
            continue
        link_key = _normalize_link(str(row[5] or ""))
        if link_key and link_key in seen_links:
            continue
        if link_key:
            seen_links.add(link_key)
        existing.append(row)

    # Fetch & filter feeds. FEEDS requires a pathogen match (Category=
    # Outbreak); OFFICIAL_FEEDS does not (Category=Regulatory) — see the
    # module docstring / config section for why the two groups differ.
    outbreak_rows, outbreak_per_feed = _process_feed_group(
        FEEDS, cutoff, seen_links, retrieved, require_pathogen=True, category="Outbreak")
    official_rows, official_per_feed = _process_feed_group(
        OFFICIAL_FEEDS, cutoff, seen_links, retrieved, require_pathogen=False, category="Regulatory")
    existing.extend(outbreak_rows)
    existing.extend(official_rows)
    added = len(outbreak_rows) + len(official_rows)
    per_feed_added: dict[str, int] = {**outbreak_per_feed, **official_per_feed}

    # Sort newest-first by Published (UTC).
    def _sort_key(row):
        dt = _parse_utc(row[0])
        return dt or datetime.min.replace(tzinfo=timezone.utc)
    existing.sort(key=_sort_key, reverse=True)

    # Rewrite NEWS sheet: clear, header, rows.
    ws.delete_rows(1, ws.max_row)
    ws.append(NEWS_HEADERS)
    for row in existing:
        ws.append(row)

    wb.save(xlsx_path)

    print(f"[news] done: +{added} new, -{dropped_old} aged out, {len(existing)} total")
    for src, n in sorted(per_feed_added.items(), key=lambda x: -x[1]):
        print(f"        {src}: +{n}")

    # Mirror NEWS sheet -> docs/data/news.json (public-facing export, same
    # sanctioned pattern as mirror_json_from_xlsx() for recalls.json).
    # Non-fatal: the xlsx write above is the critical operation, the JSON
    # mirror is a nice-to-have the dashboard reads.
    try:
        root = Path(__file__).resolve().parent.parent
        if str(root) not in sys.path:
            sys.path.insert(0, str(root))
        from pipeline.merge_master import mirror_news_json_from_xlsx
        n_written = mirror_news_json_from_xlsx(xlsx_path, root / "docs" / "data" / "news.json")
        print(f"[news] mirrored {n_written} rows -> docs/data/news.json")
    except Exception as e:
        print(f"[news] WARNING: news.json mirror failed (non-fatal): {e}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
